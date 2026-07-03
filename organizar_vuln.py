import pandas as pd
from datetime import datetime, timedelta
from colorama import Fore, init
import os

# Importações para formatar o Excel
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# Inicializa as cores do terminal
init(autoreset=True)

# --- CONFIGURAÇÕES ---
ARQUIVO_ORIGEM = "vulnerabilidades_semestre.xlsx"
ARQUIVO_DESTINO = "relatorio_quinzenal_final.xlsx"

PESO_ATIVOS = {
    "Windows Server": 5,
    "Elasticsearch": 4,
    "Aruba": 3,
    "ChromeOS": 2,
    "Tor Browser": 1
}

PESO_CRITICIDADE = {
    "Crítico": 1.5, "Alto": 1.3, "Médio": 1.1, "Baixo": 1.0, "N/A": 1.0
}


def gerar_relatorio_final():
    # --- CORREÇÃO 1: Verifica se o arquivo de origem existe ---
    if not os.path.exists(ARQUIVO_ORIGEM):
        print(f"{Fore.RED}❌ Base de dados '{ARQUIVO_ORIGEM}' não encontrada.")
        return

    try:
        print(f"{Fore.BLUE}🚀 Iniciando Triagem Contextual de Risco Avançada...")
        df = pd.read_excel(ARQUIVO_ORIGEM)

        # --- CORREÇÃO 2: Verifica colunas obrigatórias antes de prosseguir ---
        colunas_obrigatorias = ["Data Publicação CVE", "CVSS Score", "Ativo", "Criticidade", "CVE"]
        faltando = [c for c in colunas_obrigatorias if c not in df.columns]
        if faltando:
            print(f"{Fore.RED}❌ Colunas ausentes na planilha: {faltando}")
            return

        # 1. Tratamento Robusto de Datas
        # --- CORREÇÃO 3: dt.tz_localize(None) falha se a coluna já for tz-naive;
        #     usando tz_convert + tz_localize de forma segura ---
        datas = pd.to_datetime(df["Data Publicação CVE"], errors='coerce')
        if datas.dt.tz is not None:
            datas = datas.dt.tz_convert(None)
        df["Data Publicação CVE"] = datas
        df = df.dropna(subset=["Data Publicação CVE"])

        # 2. Filtro Quinzenal
        limite_15_dias = datetime.now() - timedelta(days=15)
        df_quinzenal = df[df["Data Publicação CVE"] >= limite_15_dias].copy()

        if df_quinzenal.empty:
            print(f"{Fore.YELLOW}☕ Nenhuma vulnerabilidade relevante nos últimos 15 dias.")
            return

        # 3. Engenharia de Atributos e Detecção Contextual
        df_quinzenal["CVSS_Num"] = pd.to_numeric(df_quinzenal["CVSS Score"], errors='coerce').fillna(0)

        # Garante que a coluna Descrição existe antes de checar palavras-chave
        if "Descrição" not in df_quinzenal.columns:
            df_quinzenal["Descrição"] = ""

        # --- CORREÇÃO 4: Garante que a coluna é string antes do .str.contains ---
        df_quinzenal["Descrição"] = df_quinzenal["Descrição"].astype(str).fillna("")

        # Detecção de RCE
        padrao_rce = r"remote|rce|code execution|arbitrary|command execution|network"
        df_quinzenal["Exposto_RCE"] = df_quinzenal["Descrição"].str.contains(
            padrao_rce, case=False, na=False
        )
        df_quinzenal["Peso_Exp"] = df_quinzenal["Exposto_RCE"].map({True: 3, False: 1})

        # --- DETECÇÃO EM TRÊS NÍVEIS ---

        # Nível 1 — KEV CISA (fonte oficial, peso 1.6)
        # Quando o 1hora.py coleta dados, grava a coluna "KEV" = "Sim"/"Não" diretamente
        # da API do NVD (campo exploitAdd). É a fonte mais confiável: curada por analistas.
        # Fallback para "Não" em planilhas antigas que ainda não têm essa coluna.
        if "KEV" not in df_quinzenal.columns:
            df_quinzenal["KEV"] = "Não"
        kev_confirmado = df_quinzenal["KEV"].str.upper() == "SIM"

        # Nível 2 — Regex em texto (fallback para CVEs sem coluna KEV ou dados antigos)
        # Peso 1.5 para sinais fortes de exploit confirmado, 1.2 para classe exploitável.
        padrao_exploit_confirmado = (
            r"(exploit(ed|ing|s)?)"                      # exploit, exploited, exploiting, exploits
            r"|(weaponized)"                             # weaponizado
            r"|(in the wild)"                            # uso real confirmado
            r"|(active(ly)? exploit)"                    # actively exploited
            r"|(proof.of.concept|poc)"                   # PoC público
            r"|(publicly (available|disclosed) exploit)" # exploit disponível publicamente
            r"|(known exploit)"                          # exploit conhecido
            r"|(0.?day|zero.?day)"                       # zero-day
            r"|(metasploit)"                             # módulo no Metasploit = exploit público
            r"|(nation.?state)"                          # ataque de estado-nação
        )
        padrao_exploravel = (
            r"(sandbox escape)"           # fuga de sandbox
            r"|(arbitrary code)"          # execução de código arbitrário
            r"|(privilege escalation)"    # escalada de privilégio
            r"|(heap buffer overflow)"    # overflow clássico de heap
            r"|(integer overflow)"        # overflow de inteiro
            r"|(type confusion)"          # confusão de tipos (V8, JIT etc.)
            r"|(use.after.free)"          # use-after-free
            r"|(out of bounds write)"     # escrita fora dos limites
            r"|(memory corruption)"       # corrupção de memória
            r"|(code injection)"          # injeção de código
            r"|(command injection)"       # injeção de comando
            r"|(buffer overflow)"         # overflow genérico
        )
        exploit_texto = df_quinzenal["Descrição"].str.contains(
            padrao_exploit_confirmado, case=False, na=False
        )
        alta_explorabilidade = df_quinzenal["Descrição"].str.contains(
            padrao_exploravel, case=False, na=False
        )

        # Aplicação em cascata: KEV > texto confirmado > alta explorabilidade > sem indicador
        df_quinzenal["Exploitavel"] = "Não"
        df_quinzenal.loc[alta_explorabilidade, "Exploitavel"] = "Alta Explorabilidade"
        df_quinzenal.loc[exploit_texto,        "Exploitavel"] = "Confirmado (texto)"
        df_quinzenal.loc[kev_confirmado,       "Exploitavel"] = "KEV CISA"  # sobrescreve tudo

        # Peso na fórmula de risco
        df_quinzenal["Peso_Exploit"] = 1.0
        df_quinzenal.loc[alta_explorabilidade, "Peso_Exploit"] = 1.2
        df_quinzenal.loc[exploit_texto,        "Peso_Exploit"] = 1.5
        df_quinzenal.loc[kev_confirmado,       "Peso_Exploit"] = 1.6  # maior que tudo

        # Mapeamento de Pesos Base
        df_quinzenal["Peso_Ativo"] = df_quinzenal["Ativo"].map(PESO_ATIVOS).fillna(2)
        df_quinzenal["Peso_Crit"] = df_quinzenal["Criticidade"].map(PESO_CRITICIDADE).fillna(1.0)

        # Cálculo de Aging Controlado
        df_quinzenal["Dias_Publicado"] = (
            datetime.now() - df_quinzenal["Data Publicação CVE"]
        ).dt.days.clip(lower=0)
        df_quinzenal["Fator_Aging"] = (1 + df_quinzenal["Dias_Publicado"] / 100).clip(upper=2)

        # 4. FÓRMULA DE RISCO REAL
        df_quinzenal["Score_Final"] = (
            df_quinzenal["CVSS_Num"]
            * (1 + df_quinzenal["Peso_Ativo"] / 5)
            * (1 + (df_quinzenal["Peso_Crit"] - 1) / 2)
            * (1 + (df_quinzenal["Peso_Exp"] - 1) / 5)
            * df_quinzenal["Peso_Exploit"]
            * df_quinzenal["Fator_Aging"]
        )

        # Ordenação com desempate por data
        df_quinzenal = df_quinzenal.sort_values(
            by=["Score_Final", "Data Publicação CVE"], ascending=[False, False]
        )

        # 5. Interface de Saída Visual
        print(f"\n{Fore.CYAN}{'='*75}")
        print(f"{Fore.CYAN}📊 SOC TRIAGE: PRIORIZAÇÃO BASEADA EM RISCO REAL (V3)")
        print(f"{Fore.CYAN}{'='*75}")

        ameaca_principal = df_quinzenal.iloc[0]

        # Tratamento de Descrição Nula
        desc = ameaca_principal.get("Descrição", "")
        desc_limpa = str(desc) if pd.notnull(desc) else "Sem descrição fornecida."
        desc_limpa = desc_limpa[:150]

        nivel_exploit = ameaca_principal["Exploitavel"]
        if nivel_exploit == "Confirmado":
            cor_exploit  = Fore.RED
            icone_exploit = "EXPLOIT CONFIRMADO"
        elif nivel_exploit == "Alta Explorabilidade":
            cor_exploit  = Fore.YELLOW
            icone_exploit = "ALTA EXPLORABILIDADE"
        else:
            cor_exploit  = Fore.GREEN
            icone_exploit = "Sem indicador"

        print(f"{Fore.RED}PRIORIDADE MAXIMA (TOP INCIDENT)")
        print(f"{Fore.WHITE}CVE:         {Fore.RED}{ameaca_principal['CVE']}")
        print(
            f"{Fore.WHITE}Ativo:       {Fore.YELLOW}{ameaca_principal['Ativo']} "
            f"(RCE: {ameaca_principal['Exposto_RCE']} | Exploit: {cor_exploit}{icone_exploit}{Fore.WHITE})"
        )
        print(f"{Fore.WHITE}AGING:       {Fore.BLUE}{ameaca_principal['Dias_Publicado']} dias online")
        print(f"{Fore.WHITE}RISCO REAL:  {Fore.MAGENTA}{ameaca_principal['Score_Final']:.2f}")
        print(f"{Fore.WHITE}Resumo:      {desc_limpa}...\n")

        print(f"{Fore.CYAN}TOP 5 AMEACAS PARA REMEDIACAO:")
        for i, (_, row) in enumerate(df_quinzenal.head(5).iterrows(), 1):
            nivel = row["Exploitavel"]
            if nivel == "Confirmado":
                alerta = f"{Fore.RED}EXPLOIT CONFIRMADO"
            elif nivel == "Alta Explorabilidade":
                alerta = f"{Fore.YELLOW}ALTA EXPLORABILIDADE"
            else:
                alerta = ""
            print(
                f"{Fore.WHITE}{i}. {row['CVE']:<15} | {row['Ativo']:<15} "
                f"| {row['Score_Final']:.2f} {alerta}"
            )

        # 6. Exportação — remove colunas auxiliares de cálculo
        colunas_remover = [
            "CVSS_Num", "Peso_Crit", "Peso_Exp", "Peso_Exploit",
            "Fator_Aging", "Exposto_RCE", "Dias_Publicado"
        ]
        df_final = df_quinzenal.drop(
            columns=[c for c in colunas_remover if c in df_quinzenal.columns]
        )
        df_final.to_excel(ARQUIVO_DESTINO, index=False)

        # ---------------------------------------------------------
        # 7. FORMATAÇÃO VISUAL DO EXCEL (OPENPYXL)
        # ---------------------------------------------------------
        print(f"{Fore.BLUE}🎨 Aplicando formatação visual na planilha...")

        wb = load_workbook(ARQUIVO_DESTINO)
        ws = wb.active

        # Paleta de cores
        cor_cabecalho = PatternFill(start_color="203764", end_color="203764", fill_type="solid")
        cor_top1      = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
        cor_top2      = PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid")
        cor_top3      = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
        cor_zebrada   = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

        borda_fina = Border(
            left=Side(style="thin", color="A6A6A6"),
            right=Side(style="thin", color="A6A6A6"),
            top=Side(style="thin", color="A6A6A6"),
            bottom=Side(style="thin", color="A6A6A6"),
        )

        # Cabeçalho (linha 1)
        for cell in ws[1]:
            cell.font      = Font(bold=True, color="FFFFFF")
            cell.fill      = cor_cabecalho
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = borda_fina

        # Linhas de dados (a partir da linha 2)
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            for cell in row:
                cell.border    = borda_fina
                cell.alignment = Alignment(vertical="center")

                if row_idx == 2:
                    cell.fill = cor_top1
                elif row_idx == 3:
                    cell.fill = cor_top2
                elif row_idx == 4:
                    cell.fill = cor_top3
                elif row_idx % 2 == 0:
                    cell.fill = cor_zebrada

        # --- CORREÇÃO 5: Ajuste de largura com tratamento de célula mesclada ---
        for col in ws.columns:
            max_length  = 0
            col_letter  = col[0].column_letter
            for cell in col:
                # Células mescladas não têm .value acessível diretamente; ignora-as
                try:
                    if cell.value is not None and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_length + 2, 60)

        wb.save(ARQUIVO_DESTINO)
        # ---------------------------------------------------------

        print(f"\n{Fore.GREEN}✅ Relatório finalizado e formatado: {ARQUIVO_DESTINO}")
        print(f"{Fore.CYAN}{'='*75}\n")

    except Exception as e:
        print(f"{Fore.RED}❌ Erro crítico: {e}")
        raise  # --- CORREÇÃO 6: re-lança para não engolir o traceback em dev ---


if __name__ == "__main__":
    gerar_relatorio_final()